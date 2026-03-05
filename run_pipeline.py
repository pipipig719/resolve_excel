from __future__ import annotations

import argparse
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

CN_CODE = "\u996e\u7247\u7f16\u7801"
CN_BATCH = "\u6279\u6b21"
CN_STOCK = "\u5e93\u5b58"
CN_STATUS = "\u72b6\u6001"
CN_ENABLED = "\u662f\u5426\u542f\u7528"
CN_LOCATION = "\u8d27\u4f4d\u7f16\u53f7"
CN_MIN_STOCK = "\u5e93\u5b58\u4e0b\u9650\u503c"

FINAL_OUTPUT_NAME = "\u6000\u5b81\u996e\u7247\u8d27\u4f4d\u5bfc\u5165\u6700\u7ec8\u6587\u4ef6.xlsx"
BACKUP_OUTPUT_NAME = "\u6000\u5b81\u996e\u7247\u8d27\u4f4d\u5bfc\u5165\u5907\u4efd\u6587\u4ef6.xlsx"

INVENTORY_HEADERS = {CN_CODE, CN_BATCH, CN_STOCK, CN_STATUS}
TEMPLATE_HEADERS = {CN_CODE, CN_ENABLED, CN_LOCATION, CN_STOCK, CN_MIN_STOCK}
SOURCE_HEADERS = {CN_CODE, CN_STOCK}

CN_NAME_INVENTORY = "\u5e93\u5b58"
CN_NAME_TEMPLATE = "\u6a21\u677f"
CN_NAME_BACKUP = "\u5907\u4efd\u6587\u4ef6"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_xlsx_workbook(path: Path) -> openpyxl.Workbook:
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except InvalidFileException:
        # Some files are xlsx content with .xls suffix.
        with path.open("rb") as fh:
            return openpyxl.load_workbook(fh, data_only=True)


def header_set_xlsx(file_path: Path) -> set[str]:
    wb = load_xlsx_workbook(file_path)
    ws = wb.worksheets[0]
    headers: set[str] = set()
    for col in range(1, ws.max_column + 1):
        key = normalize_text(ws.cell(1, col).value)
        if key:
            headers.add(key)
    return headers


def header_set_xls(file_path: Path) -> set[str]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("读取 .xls 需要 xlrd，请先执行 uv sync 安装依赖。") from exc

    try:
        wb = xlrd.open_workbook(str(file_path))
    except Exception as exc:  # noqa: BLE001
        if "xlsx file; not supported" in str(exc).lower():
            return header_set_xlsx(file_path)
        raise

    ws = wb.sheet_by_index(0)
    headers: set[str] = set()
    for col in range(ws.ncols):
        key = normalize_text(ws.cell_value(0, col))
        if key:
            headers.add(key)
    return headers


def header_set(file_path: Path) -> set[str]:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return header_set_xlsx(file_path)
    if suffix == ".xls":
        return header_set_xls(file_path)
    return set()


def looks_like_headerless_inventory(file_path: Path) -> bool:
    try:
        wb = load_xlsx_workbook(file_path)
    except Exception:  # noqa: BLE001
        return False
    ws = wb.worksheets[0]
    # Legacy inventory layout has at least 12 columns with data starting from row 1.
    return ws.max_column >= 12 and ws.max_row >= 1


def detect_inventory(root: Path, explicit: Path | None) -> Path:
    if explicit is not None:
        return explicit.resolve()

    header_candidates: list[Path] = []
    fallback_candidates: list[Path] = []

    for file in root.glob("*.xlsx"):
        if file.name.startswith("~$"):
            continue
        if file.name == FINAL_OUTPUT_NAME:
            continue

        headers = header_set(file)
        if INVENTORY_HEADERS.issubset(headers):
            header_candidates.append(file)
            continue

        if looks_like_headerless_inventory(file):
            fallback_candidates.append(file)

    if header_candidates:
        header_candidates.sort(
            key=lambda p: ((CN_NAME_INVENTORY in p.name), p.stat().st_mtime),
            reverse=True,
        )
        return header_candidates[0].resolve()

    if fallback_candidates:
        fallback_candidates.sort(
            key=lambda p: ((CN_NAME_INVENTORY in p.name), p.stat().st_mtime),
            reverse=True,
        )
        return fallback_candidates[0].resolve()

    raise FileNotFoundError("未找到库存文件（需包含表头，或旧版无表头库存结构）")


def detect_template(root: Path, explicit: Path | None) -> Path:
    if explicit is not None:
        return explicit.resolve()

    candidates: list[Path] = []
    for file in root.glob("*.xlsx"):
        if file.name.startswith("~$"):
            continue
        headers = header_set(file)
        if TEMPLATE_HEADERS.issubset(headers):
            candidates.append(file)

    if not candidates:
        raise FileNotFoundError("未找到模板文件（需包含：饮片编码、是否启用、货位编号、库存、库存下限值）")

    preferred = [p for p in candidates if CN_NAME_TEMPLATE in p.name]
    if preferred:
        preferred.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return preferred[0].resolve()

    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0].resolve()


def detect_source_inventory(source_dir: Path, explicit: Path | None) -> Path:
    if explicit is not None:
        return explicit.resolve()

    if not source_dir.exists():
        raise FileNotFoundError(f"source 目录不存在: {source_dir}")

    candidates: list[Path] = []
    xls_without_xlrd = False

    for file in source_dir.iterdir():
        if not file.is_file():
            continue
        if file.name.startswith("~$"):
            continue
        if file.suffix.lower() not in {".xlsx", ".xls"}:
            continue
        if CN_NAME_BACKUP in file.name:
            continue
        try:
            headers = header_set(file)
        except RuntimeError:
            if file.suffix.lower() == ".xls":
                xls_without_xlrd = True
                continue
            raise

        if SOURCE_HEADERS.issubset(headers):
            candidates.append(file)

    if not candidates:
        if xls_without_xlrd:
            raise RuntimeError("检测到 .xls 源文件，但缺少 xlrd。请先执行 uv sync。")
        raise FileNotFoundError("source 下未找到配方间库存文件（需包含：饮片编码、库存）")

    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0].resolve()


def run_cmd(command: list[str], cwd: Path) -> None:
    result = subprocess.run(command, cwd=cwd, text=True)
    if result.returncode != 0:
        raise RuntimeError(f"命令执行失败（exit={result.returncode}）: {' '.join(command)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="一键生成：根目录最终导入文件 + source 目录备份文件。"
    )
    parser.add_argument("--inventory", help="库存文件路径（可选，默认自动识别）")
    parser.add_argument("--template", help="模板文件路径（可选，默认自动识别）")
    parser.add_argument("--source-file", help="source 下配方间数据文件（可选，默认自动识别）")
    parser.add_argument("--final-output", default=FINAL_OUTPUT_NAME, help="根目录最终文件名")
    parser.add_argument(
        "--backup-output",
        default=f"source/{BACKUP_OUTPUT_NAME}",
        help="source 目录备份文件路径",
    )
    parser.add_argument("--keep-reports", action="store_true", help="保留中间报告文件")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = Path(__file__).resolve().parent
    source_dir = root / "source"

    inventory_path = detect_inventory(root, Path(args.inventory) if args.inventory else None)
    template_path = detect_template(root, Path(args.template) if args.template else None)
    source_path = detect_source_inventory(source_dir, Path(args.source_file) if args.source_file else None)

    final_output = (root / args.final_output).resolve()
    backup_output = (root / args.backup_output).resolve()
    report_dir = (root / ".tmp_reports").resolve()

    print("[INFO] 检测到文件：")
    print(f"inventory = {inventory_path}")
    print(f"template  = {template_path}")
    print(f"source    = {source_path}")
    print(f"final_out = {final_output}")
    print(f"backup_out= {backup_output}")

    run_cmd(
        [
            sys.executable,
            "convert_inventory.py",
            "--inventory",
            str(inventory_path),
            "--template",
            str(template_path),
            "--source-for-match",
            str(source_path),
            "--output",
            str(final_output),
            "--report-dir",
            str(report_dir),
        ],
        cwd=root,
    )

    run_cmd(
        [
            sys.executable,
            "convert_source_backup.py",
            "--source",
            str(source_path),
            "--template",
            str(template_path),
            "--output",
            str(backup_output),
        ],
        cwd=root,
    )

    if not args.keep_reports and report_dir.exists():
        shutil.rmtree(report_dir, ignore_errors=True)

    print("[OK] 全部完成。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
