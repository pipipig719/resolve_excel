from __future__ import annotations

import argparse
import re
import sys
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

CN_CODE = "\u996e\u7247\u7f16\u7801"
CN_STOCK = "\u5e93\u5b58"
CN_ENABLED = "\u662f\u5426\u542f\u7528"
CN_LOCATION = "\u8d27\u4f4d\u7f16\u53f7"
CN_MIN_STOCK = "\u5e93\u5b58\u4e0b\u9650\u503c"

CN_STATUS_ON = "\u542f\u7528"
CN_STATUS_OFF = "\u7981\u7528"
CN_YES = "\u662f"
CN_NO = "\u5426"

SOURCE_REQUIRED_HEADERS = (CN_CODE, CN_STOCK)
TEMPLATE_REQUIRED_HEADERS = (CN_CODE, CN_ENABLED, CN_LOCATION, CN_STOCK, CN_MIN_STOCK)
RE_NUMERIC_TEXT = re.compile(r"^[+-]?(?:\d+(?:\.\d+)?|\.\d+)(?:[eE][+-]?\d+)?$")


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def code_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return _normalize_numeric_string(str(value))

    text = normalize_text(value)
    if text == "":
        return ""

    plain = text.replace(",", "")
    if re.match(r"^[+-]?0\d+$", plain):
        return text
    if not RE_NUMERIC_TEXT.match(plain):
        return text
    return _normalize_numeric_string(plain)


def _parse_decimal(raw: str) -> Decimal | None:
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _normalize_numeric_string(raw: str) -> str:
    dec = _parse_decimal(raw)
    if dec is None:
        return raw
    if dec == dec.to_integral_value():
        return format(dec.quantize(Decimal("1")), "f")
    normalized = dec.normalize()
    text = format(normalized, "f")
    return text.rstrip("0").rstrip(".")


def to_number(value: object) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        dec = Decimal(str(value))
        int_part = dec.to_integral_value()
        if dec == int_part:
            return int(int_part)
        return float(dec.quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP))
    raw = normalize_text(value).replace(",", "")
    if raw == "":
        return None
    dec = _parse_decimal(raw)
    if dec is None:
        return value
    int_part = dec.to_integral_value()
    if dec == int_part:
        return int(int_part)
    return float(dec.quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP))


def map_enabled(raw: object) -> str:
    text = normalize_text(raw).lower()
    if text in {CN_STATUS_ON, CN_YES, "1", "true"}:
        return CN_YES
    if text in {CN_STATUS_OFF, CN_NO, "0", "false"}:
        return CN_NO
    return normalize_text(raw)


def read_header_map_xlsx(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = normalize_text(ws.cell(1, col).value)
        if key and key not in header_map:
            header_map[key] = col
    return header_map


def read_source_rows_xlsx(source_path: Path) -> tuple[list[dict[str, object]], dict[str, int]]:
    try:
        wb = openpyxl.load_workbook(source_path, data_only=True)
    except InvalidFileException:
        # Handle xlsx-content files that use a .xls suffix.
        with source_path.open("rb") as fh:
            wb = openpyxl.load_workbook(fh, data_only=True)
    ws = wb.worksheets[0]
    headers = read_header_map_xlsx(ws)
    rows: list[dict[str, object]] = []
    for row_idx in range(2, ws.max_row + 1):
        row: dict[str, object] = {}
        for name, col in headers.items():
            row[name] = ws.cell(row_idx, col).value
        rows.append(row)
    return rows, headers


def read_source_rows_xls(source_path: Path) -> tuple[list[dict[str, object]], dict[str, int]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("读取 .xls 需要 xlrd，请先执行 uv sync 安装依赖。") from exc

    try:
        wb = xlrd.open_workbook(str(source_path))
    except Exception as exc:  # noqa: BLE001
        # Some files are xlsx content with .xls suffix.
        if "xlsx file; not supported" in str(exc).lower():
            return read_source_rows_xlsx(source_path)
        raise
    ws = wb.sheet_by_index(0)
    headers: dict[str, int] = {}
    for col in range(ws.ncols):
        key = normalize_text(ws.cell_value(0, col))
        if key and key not in headers:
            headers[key] = col

    rows: list[dict[str, object]] = []
    for row_idx in range(1, ws.nrows):
        row: dict[str, object] = {}
        for name, col in headers.items():
            row[name] = ws.cell_value(row_idx, col)
        rows.append(row)
    return rows, headers


def read_source_rows(source_path: Path) -> tuple[list[dict[str, object]], dict[str, int]]:
    suffix = source_path.suffix.lower()
    if suffix == ".xlsx":
        return read_source_rows_xlsx(source_path)
    if suffix == ".xls":
        return read_source_rows_xls(source_path)
    raise ValueError(f"不支持的源文件格式: {source_path.suffix}")


def ensure_headers(header_map: dict[str, int], required: tuple[str, ...], label: str) -> None:
    missing = [name for name in required if name not in header_map]
    if missing:
        raise ValueError(f"{label} 缺少列: {', '.join(missing)}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="把配方间源数据转换成导入备份文件，输出格式严格跟随模板。"
    )
    parser.add_argument("--source", required=True, help="配方间源数据文件（.xlsx 或 .xls）")
    parser.add_argument("--template", required=True, help="导入模板文件（.xlsx）")
    parser.add_argument("--output", required=True, help="输出备份文件（.xlsx）")
    parser.add_argument("--default-location", default="Z999", help="缺失货位时默认值")
    parser.add_argument("--default-min-stock", type=float, default=500, help="缺失库存下限时默认值")
    parser.add_argument("--sort-desc", action="store_true", help="按库存从高到低排序")
    return parser


def main() -> int:
    args = build_parser().parse_args()

    source_path = Path(args.source).expanduser().resolve()
    template_path = Path(args.template).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    if not source_path.exists():
        print(f"[ERROR] 源文件不存在: {source_path}", file=sys.stderr)
        return 1
    if not template_path.exists():
        print(f"[ERROR] 模板不存在: {template_path}", file=sys.stderr)
        return 1

    try:
        src_rows, src_headers = read_source_rows(source_path)
        ensure_headers(src_headers, SOURCE_REQUIRED_HEADERS, f"源文件 {source_path}")

        tpl_wb = openpyxl.load_workbook(template_path)
        tpl_ws = tpl_wb.worksheets[0]
        tpl_headers = read_header_map_xlsx(tpl_ws)
        ensure_headers(tpl_headers, TEMPLATE_REQUIRED_HEADERS, f"模板 {template_path}")

        rows: list[tuple[str, str, str, float | int | None, float | int | None]] = []

        default_min = args.default_min_stock
        default_min_value: float | int
        if abs(default_min - round(default_min)) < 1e-9:
            default_min_value = int(round(default_min))
        else:
            default_min_value = round(default_min, 6)

        for src_row in src_rows:
            code = code_to_text(src_row.get(CN_CODE))
            if not code:
                continue

            enabled = map_enabled(src_row.get(CN_ENABLED, ""))
            location = normalize_text(src_row.get(CN_LOCATION, "")) or args.default_location
            stock = to_number(src_row.get(CN_STOCK))
            min_stock = to_number(src_row.get(CN_MIN_STOCK))
            if min_stock is None:
                min_stock = default_min_value

            rows.append((code, enabled, location, stock, min_stock))

        if args.sort_desc:
            rows.sort(
                key=lambda item: float(item[3]) if item[3] is not None else float("-inf"),
                reverse=True,
            )

        code_col = tpl_headers[CN_CODE]
        enabled_col = tpl_headers[CN_ENABLED]
        location_col = tpl_headers[CN_LOCATION]
        stock_col = tpl_headers[CN_STOCK]
        min_stock_col = tpl_headers[CN_MIN_STOCK]

        if tpl_ws.max_row >= 2:
            tpl_ws.delete_rows(2, tpl_ws.max_row - 1)

        for out_row, row_data in enumerate(rows, start=2):
            code_cell = tpl_ws.cell(out_row, code_col, row_data[0])
            enabled_cell = tpl_ws.cell(out_row, enabled_col, row_data[1])
            location_cell = tpl_ws.cell(out_row, location_col, row_data[2])
            stock_cell = tpl_ws.cell(out_row, stock_col, row_data[3])
            min_cell = tpl_ws.cell(out_row, min_stock_col, row_data[4])

            code_cell.number_format = "@"
            enabled_cell.number_format = "@"
            location_cell.number_format = "@"
            stock_cell.number_format = "0.######"
            min_cell.number_format = "0.######"

        output_path.parent.mkdir(parents=True, exist_ok=True)
        tpl_wb.save(output_path)
    except Exception as exc:  # noqa: BLE001
        print(f"[ERROR] 处理失败: {exc}", file=sys.stderr)
        return 1

    print("[OK] 备份模板生成完成")
    print(f"source:   {source_path}")
    print(f"template: {template_path}")
    print(f"output:   {output_path}")
    print(f"rows:     {len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
