from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

CN_CODE = "\u6807\u8bc6\u7801"
CN_NAME = "\u54c1\u540d"
CN_SPEC = "\u89c4\u683c"
CN_STOCK_INT = "\u5e93\u5b58\u6570|\u6574\u6570"
CN_STOCK_REM = "\u5e93\u5b58\u6570|\u4f59\u6570"
CN_MIN = "\u5e93\u5b58\u4e0b\u9650"
CN_SYS_DISABLE = "\u7cfb\u7edf\u505c\u7528"

T_CODE = "\u996e\u7247\u7f16\u7801"
T_ENABLED = "\u662f\u5426\u542f\u7528"
T_LOCATION = "\u8d27\u4f4d\u7f16\u53f7"
T_STOCK = "\u5e93\u5b58"
T_MIN = "\u5e93\u5b58\u4e0b\u9650\u503c"

S_CODE = "\u996e\u7247\u7f16\u7801"
S_STOCK = "\u5e93\u5b58"
S_LOCATION = "\u8d27\u4f4d\u7f16\u53f7"
S_MIN = "\u5e93\u5b58\u4e0b\u9650\u503c"
S_ENABLED = "\u662f\u5426\u542f\u7528"

CN_YES = "\u662f"
CN_NO = "\u5426"
CN_ENABLE = "\u542f\u7528"
CN_DISABLE = "\u7981\u7528"

RE_NUMBER_UNIT = re.compile(
    r"^\s*([+-]?(?:\d+(?:\.\d+)?|\.\d+)(?:[eE][+-]?\d+)?)\s*([A-Za-z\u4e00-\u9fff]*)\s*$"
)
RE_NUMERIC_TEXT = re.compile(r"^[+-]?(?:\d+(?:\.\d+)?|\.\d+)(?:[eE][+-]?\d+)?$")


@dataclass
class InvRow:
    row: int
    code: str
    name: str
    spec: str
    int_raw: str
    rem_raw: str
    min_raw: str
    disable_raw: str
    stock_value: float
    min_value: float | None
    enabled: str


def decode_maybe_gbk(value: object) -> object:
    if isinstance(value, str):
        try:
            return value.encode("latin1").decode("gbk")
        except Exception:
            return value
    return value


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = decode_maybe_gbk(value)
    return str(text).strip()


def _parse_decimal(raw: str) -> Decimal | None:
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _normalize_numeric_string(raw: str) -> str:
    dec = _parse_decimal(raw)
    if dec is None:
        return raw
    if dec == dec.to_integral_value():
        return format(dec.quantize(Decimal("1")), "f")
    normalized = dec.normalize()
    text = format(normalized, "f")
    return text.rstrip("0").rstrip(".")


def code_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return _normalize_numeric_string(str(value))

    text = normalize_text(value)
    if text == "":
        return ""
    plain = text.replace(",", "")
    if re.match(r"^[+-]?0\d+$", plain):
        return text
    if not RE_NUMERIC_TEXT.match(plain):
        return text
    return _normalize_numeric_string(plain)


def parse_number_with_unit(raw: object) -> tuple[float, str]:
    text = normalize_text(raw)
    if text == "":
        return 0.0, ""
    m = RE_NUMBER_UNIT.match(text)
    if not m:
        return 0.0, ""
    num_dec = _parse_decimal(m.group(1))
    if num_dec is None:
        return 0.0, ""
    num = float(num_dec)
    unit = m.group(2).strip()
    return num, unit


def to_stock_amount(raw: object) -> float:
    num, unit = parse_number_with_unit(raw)
    unit_l = unit.lower()
    if unit_l in {"kg"} or unit in {"\u516c\u65a4"}:
        return num * 1000.0
    # For g or other units (\u6761/\u4e2a/\u5305), keep numeric value as-is.
    return num


def parse_min_value(raw: object) -> float | None:
    text = normalize_text(raw)
    if text == "":
        return None
    num, unit = parse_number_with_unit(text)
    if unit.lower() in {"kg"} or unit in {"\u516c\u65a4"}:
        return num * 1000.0
    return num


def map_enabled_from_disable(raw: object) -> str:
    text = normalize_text(raw).lower()
    if text in {"", "0", "\u5426", "false", CN_ENABLE.lower()}:
        return CN_YES
    if text in {"1", "\u662f", "true", CN_DISABLE.lower(), "\u505c\u7528"}:
        return CN_NO
    if "\u505c\u7528" in text or "\u7981\u7528" in text:
        return CN_NO
    return CN_YES


def excel_num(value: float) -> float | int:
    dec = Decimal(str(value))
    int_part = dec.to_integral_value()
    if dec == int_part:
        return int(int_part)
    return float(dec.quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP))


def find_first_file(directory: Path, exts: tuple[str, ...], exclude_keywords: tuple[str, ...] = ()) -> Path:
    candidates = []
    for p in directory.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() not in exts:
            continue
        if any(k in p.name for k in exclude_keywords):
            continue
        candidates.append(p)
    if not candidates:
        raise FileNotFoundError(f"No file found in {directory} with extensions {exts}")
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0]


def find_template_file(project_root: Path) -> Path:
    files = [p for p in project_root.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")]
    preferred = [p for p in files if "\u6a21\u677f" in p.name]
    if preferred:
        preferred.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        return preferred[0]
    if not files:
        raise FileNotFoundError("No template xlsx found in project root")
    files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return files[0]


def _parse_inventory_rows_from_table(headers: list[str], get_row_values, row_count: int) -> list[InvRow]:
    h = {name: idx for idx, name in enumerate(headers) if name}

    required = [CN_CODE, CN_STOCK_INT, CN_STOCK_REM, CN_MIN]
    missing = [x for x in required if x not in h]
    if missing:
        raise ValueError(f"Inventory missing headers: {missing}")

    rows: list[InvRow] = []
    for r in range(1, row_count):
        row_values = get_row_values(r)
        code = code_to_text(row_values[h[CN_CODE]] if h[CN_CODE] < len(row_values) else "")
        if not code:
            continue
        name = normalize_text(row_values[h[CN_NAME]] if CN_NAME in h and h[CN_NAME] < len(row_values) else "")
        spec = normalize_text(row_values[h[CN_SPEC]] if CN_SPEC in h and h[CN_SPEC] < len(row_values) else "")
        int_raw = normalize_text(row_values[h[CN_STOCK_INT]] if h[CN_STOCK_INT] < len(row_values) else "")
        rem_raw = normalize_text(row_values[h[CN_STOCK_REM]] if h[CN_STOCK_REM] < len(row_values) else "")
        min_raw = normalize_text(row_values[h[CN_MIN]] if h[CN_MIN] < len(row_values) else "")
        disable_raw = normalize_text(
            row_values[h[CN_SYS_DISABLE]] if CN_SYS_DISABLE in h and h[CN_SYS_DISABLE] < len(row_values) else ""
        )

        stock_value = to_stock_amount(int_raw) + to_stock_amount(rem_raw)
        min_value = parse_min_value(min_raw)
        enabled = map_enabled_from_disable(disable_raw)

        rows.append(
            InvRow(
                row=r + 1,
                code=code,
                name=name,
                spec=spec,
                int_raw=int_raw,
                rem_raw=rem_raw,
                min_raw=min_raw,
                disable_raw=disable_raw,
                stock_value=stock_value,
                min_value=min_value,
                enabled=enabled,
            )
        )
    return rows


def read_inventory_rows(inventory_file: Path) -> list[InvRow]:
    suffix = inventory_file.suffix.lower()

    if suffix == ".xlsx":
        wb = load_workbook_any_excel(inventory_file)
        ws = wb.worksheets[0]
        headers = [normalize_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

        def row_values(idx: int):
            row_num = idx + 1
            return [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]

        rows = _parse_inventory_rows_from_table(headers, row_values, ws.max_row)
        if not rows:
            raise ValueError("No valid inventory rows found")
        return rows

    if suffix != ".xls":
        raise ValueError(f"Unsupported inventory format: {inventory_file.suffix}")

    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("xlrd is required for .xls inventory. Run: uv sync") from exc

    wb = xlrd.open_workbook(str(inventory_file))
    ws = wb.sheet_by_index(0)
    headers = [normalize_text(v) for v in ws.row_values(0)]

    def row_values(idx: int):
        return ws.row_values(idx)

    rows = _parse_inventory_rows_from_table(headers, row_values, ws.nrows)
    return rows


def load_workbook_any_excel(path: Path):
    # for xlsx and xls-with-xlsx-content
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except InvalidFileException:
        with path.open("rb") as fh:
            return openpyxl.load_workbook(fh, data_only=True)


def load_source_rows(source_file: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []

    suffix = source_file.suffix.lower()
    if suffix == ".xlsx":
        wb = load_workbook_any_excel(source_file)
        ws = wb.worksheets[0]
        header = {normalize_text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
        if S_CODE not in header:
            raise ValueError(f"Source missing header: {S_CODE}")
        for r in range(2, ws.max_row + 1):
            code = code_to_text(ws.cell(r, header[S_CODE]).value)
            if not code:
                continue
            rows.append(
                {
                    S_CODE: code,
                    S_STOCK: normalize_text(ws.cell(r, header[S_STOCK]).value) if S_STOCK in header else "",
                    S_LOCATION: normalize_text(ws.cell(r, header[S_LOCATION]).value) if S_LOCATION in header else "",
                    S_MIN: normalize_text(ws.cell(r, header[S_MIN]).value) if S_MIN in header else "",
                    S_ENABLED: normalize_text(ws.cell(r, header[S_ENABLED]).value) if S_ENABLED in header else "",
                }
            )
        return rows

    if suffix == ".xls":
        try:
            import xlrd

            wb = xlrd.open_workbook(str(source_file))
            ws = wb.sheet_by_index(0)
            header = {normalize_text(v): i for i, v in enumerate(ws.row_values(0))}
            if S_CODE not in header:
                raise ValueError(f"Source missing header: {S_CODE}")
            for r in range(1, ws.nrows):
                rv = ws.row_values(r)
                code = code_to_text(rv[header[S_CODE]] if header[S_CODE] < len(rv) else "")
                if not code:
                    continue
                rows.append(
                    {
                        S_CODE: code,
                        S_STOCK: normalize_text(rv[header[S_STOCK]]) if S_STOCK in header and header[S_STOCK] < len(rv) else "",
                        S_LOCATION: normalize_text(rv[header[S_LOCATION]]) if S_LOCATION in header and header[S_LOCATION] < len(rv) else "",
                        S_MIN: normalize_text(rv[header[S_MIN]]) if S_MIN in header and header[S_MIN] < len(rv) else "",
                        S_ENABLED: normalize_text(rv[header[S_ENABLED]]) if S_ENABLED in header and header[S_ENABLED] < len(rv) else "",
                    }
                )
            return rows
        except Exception:
            # fallback: xlsx content with .xls suffix
            wb = load_workbook_any_excel(source_file)
            ws = wb.worksheets[0]
            header = {normalize_text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
            if S_CODE not in header:
                raise ValueError(f"Source missing header: {S_CODE}")
            for r in range(2, ws.max_row + 1):
                code = code_to_text(ws.cell(r, header[S_CODE]).value)
                if not code:
                    continue
                rows.append(
                    {
                        S_CODE: code,
                        S_STOCK: normalize_text(ws.cell(r, header[S_STOCK]).value) if S_STOCK in header else "",
                        S_LOCATION: normalize_text(ws.cell(r, header[S_LOCATION]).value) if S_LOCATION in header else "",
                        S_MIN: normalize_text(ws.cell(r, header[S_MIN]).value) if S_MIN in header else "",
                        S_ENABLED: normalize_text(ws.cell(r, header[S_ENABLED]).value) if S_ENABLED in header else "",
                    }
                )
            return rows

    raise ValueError(f"Unsupported source format: {source_file.suffix}")


def map_source_enabled(raw: object) -> str:
    t = normalize_text(raw).lower()
    if t in {"", "0", "\u5426", "false", CN_DISABLE.lower()}:
        return CN_NO if t in {"\u5426", "0", "false", CN_DISABLE.lower()} else CN_YES
    if t in {"1", "\u662f", "true", CN_ENABLE.lower()}:
        return CN_YES
    if "\u505c\u7528" in t or "\u7981\u7528" in t:
        return CN_NO
    return CN_YES


def write_template_rows(template_file: Path, out_file: Path, rows: list[dict[str, object]]) -> None:
    wb = openpyxl.load_workbook(template_file)
    ws = wb.worksheets[0]
    h = {normalize_text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    required = [T_CODE, T_ENABLED, T_LOCATION, T_STOCK, T_MIN]
    missing = [x for x in required if x not in h]
    if missing:
        raise ValueError(f"Template missing headers: {missing}")

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    for i, row in enumerate(rows, start=2):
        code_cell = ws.cell(i, h[T_CODE], row[T_CODE])
        enabled_cell = ws.cell(i, h[T_ENABLED], row[T_ENABLED])
        location_cell = ws.cell(i, h[T_LOCATION], row[T_LOCATION])
        stock_cell = ws.cell(i, h[T_STOCK], row[T_STOCK])
        min_cell = ws.cell(i, h[T_MIN], row[T_MIN])

        code_cell.number_format = "@"
        enabled_cell.number_format = "@"
        location_cell.number_format = "@"
        stock_cell.number_format = "0.######"
        min_cell.number_format = "0.######"

    out_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_file)


def build_source_profile_map(source_rows: list[dict[str, object]]) -> dict[str, dict[str, object]]:
    result: dict[str, dict[str, object]] = {}
    for r in source_rows:
        code = code_to_text(r.get(S_CODE, ""))
        if not code:
            continue
        enabled_raw = normalize_text(r.get(S_ENABLED, ""))
        enabled = map_source_enabled(enabled_raw)
        location = normalize_text(r.get(S_LOCATION, ""))
        min_val = parse_min_value(r.get(S_MIN, ""))
        stock_val = to_stock_amount(r.get(S_STOCK, ""))

        profile = result.get(code)
        if profile is None:
            result[code] = {
                "enabled": enabled,
                "location": location,
                "min": min_val,
                "stock": stock_val,
            }
            continue

        profile["stock"] = float(profile.get("stock", 0.0)) + stock_val
        if location:
            profile["location"] = location
        if min_val is not None:
            profile["min"] = min_val
        if enabled_raw:
            profile["enabled"] = enabled
    return result


def build_final_rows(
    inv_rows: list[InvRow],
    source_profile_map: dict[str, dict[str, object]],
    default_min: float,
    default_loc: str,
) -> tuple[list[dict[str, object]], Counter[str], int, int, int]:
    code_counter = Counter(r.code for r in inv_rows)
    agg_stock: dict[str, float] = defaultdict(float)
    order: list[str] = []
    seen: set[str] = set()

    for r in inv_rows:
        agg_stock[r.code] += r.stock_value
        if r.code not in seen:
            seen.add(r.code)
            order.append(r.code)

    for code in source_profile_map.keys():
        if code not in seen:
            seen.add(code)
            order.append(code)

    out: list[dict[str, object]] = []
    matched_source_count = 0
    source_only_forced_disabled_count = 0
    inv_non_positive_forced_disabled_count = 0

    for code in order:
        source_profile = source_profile_map.get(code)
        has_inventory = code in agg_stock
        inv_stock = agg_stock.get(code, 0.0)
        if source_profile is not None:
            location = normalize_text(source_profile.get("location", "")) or default_loc
            min_from_source = source_profile.get("min")
            min_val = default_min if min_from_source is None else float(min_from_source)
            matched_source_count += 1
        else:
            location = default_loc
            min_val = default_min

        # New Feixi status rules:
        # 1) Source has code but inventory does not -> disable
        # 2) Code exists in both, and inventory stock <= 0 -> disable
        # 3) Otherwise fallback to mapped source status or default yes
        if source_profile is not None and not has_inventory:
            enabled = CN_NO
            source_only_forced_disabled_count += 1
        elif source_profile is not None and has_inventory and inv_stock <= 0:
            enabled = CN_NO
            inv_non_positive_forced_disabled_count += 1
        elif source_profile is not None:
            enabled = str(source_profile.get("enabled", CN_YES))
        else:
            enabled = CN_YES

        stock_value = inv_stock
        if not has_inventory and source_profile is not None:
            stock_value = float(source_profile.get("stock", 0.0))

        out.append(
            {
                T_CODE: code,
                T_ENABLED: enabled,
                T_LOCATION: location,
                T_STOCK: excel_num(stock_value),
                T_MIN: excel_num(min_val),
            }
        )

    out.sort(key=lambda x: float(x[T_STOCK]), reverse=True)
    return (
        out,
        code_counter,
        matched_source_count,
        source_only_forced_disabled_count,
        inv_non_positive_forced_disabled_count,
    )


def build_backup_rows(source_rows: list[dict[str, object]], default_min: float, default_loc: str) -> list[dict[str, object]]:
    out: list[dict[str, object]] = []
    for r in source_rows:
        code = code_to_text(r.get(S_CODE, ""))
        if not code:
            continue
        stock_num = to_stock_amount(r.get(S_STOCK, ""))
        min_raw = r.get(S_MIN, "")
        min_val = parse_min_value(min_raw)
        if min_val is None:
            min_val = default_min
        loc = normalize_text(r.get(S_LOCATION, "")) or default_loc
        enabled = map_source_enabled(r.get(S_ENABLED, ""))
        out.append(
            {
                T_CODE: code,
                T_ENABLED: enabled,
                T_LOCATION: loc,
                T_STOCK: excel_num(stock_num),
                T_MIN: excel_num(min_val),
            }
        )
    return out


def write_duplicate_csv(path: Path, inv_rows: list[InvRow], code_counter: Counter[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["code", "count", "row", "name", "spec", "stock_int", "stock_rem", "min_raw"])
        for r in inv_rows:
            if code_counter[r.code] > 1:
                w.writerow([r.code, code_counter[r.code], r.row, r.name, r.spec, r.int_raw, r.rem_raw, r.min_raw])


def write_non_g_report(path: Path, inv_rows: list[InvRow], source_rows: list[dict[str, object]]) -> None:
    source_stock: dict[str, str] = {}
    for r in source_rows:
        code = code_to_text(r.get(S_CODE, ""))
        if not code:
            continue
        source_stock[code] = normalize_text(r.get(S_STOCK, ""))

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "code",
                "name",
                "spec",
                "inventory_int_raw",
                "inventory_rem_raw",
                "inventory_converted_stock",
                "source_stock",
            ]
        )
        for r in inv_rows:
            if not r.spec or "g" in r.spec.lower():
                continue
            w.writerow(
                [
                    r.code,
                    r.name,
                    r.spec,
                    r.int_raw,
                    r.rem_raw,
                    excel_num(r.stock_value),
                    source_stock.get(r.code, ""),
                ]
            )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Feixi inventory converter")
    parser.add_argument("--inventory", help="Feixi inventory file (.xls)")
    parser.add_argument("--source", help="Feixi source file (.xls/.xlsx)")
    parser.add_argument("--template", help="Template xlsx file")
    parser.add_argument("--output-final", help="Final output xlsx path")
    parser.add_argument("--output-backup", help="Backup output xlsx path")
    parser.add_argument("--output-duplicates", help="Duplicate code csv path")
    parser.add_argument("--output-non-g-report", help="Non-g spec compare csv path")
    parser.add_argument("--default-min-stock", type=float, default=500)
    parser.add_argument("--default-location", default="Z999")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    feixi_dir = Path(__file__).resolve().parent
    project_root = feixi_dir.parent

    source_parent = feixi_dir / "source"
    if not source_parent.exists():
        alt = feixi_dir / "feixi_source"
        if alt.exists():
            source_parent = alt

    exclude_inventory = (
        "\u5907\u4efd\u6587\u4ef6",
        "\u6700\u7ec8\u6587\u4ef6",
        "\u6a21\u677f",
        "\u91cd\u590d\u6e05\u5355",
        "process_feixi",
        "run_feixi",
    )
    if args.inventory:
        inventory_file = Path(args.inventory).resolve()
    else:
        try:
            inventory_file = find_first_file(feixi_dir, (".xls",), exclude_inventory)
        except FileNotFoundError:
            inventory_file = find_first_file(feixi_dir, (".xlsx",), exclude_inventory)
    source_file = Path(args.source).resolve() if args.source else find_first_file(source_parent, (".xls", ".xlsx"), ("\u5907\u4efd\u6587\u4ef6",))
    template_file = Path(args.template).resolve() if args.template else find_template_file(project_root)

    final_out = Path(args.output_final).resolve() if args.output_final else (feixi_dir / "\u80a5\u897f\u996e\u7247\u8d27\u4f4d\u5bfc\u5165\u6700\u7ec8\u6587\u4ef6.xlsx")
    backup_out = Path(args.output_backup).resolve() if args.output_backup else (source_parent / "\u80a5\u897f\u996e\u7247\u8d27\u4f4d\u5bfc\u5165\u5907\u4efd\u6587\u4ef6.xlsx")
    dup_csv = Path(args.output_duplicates).resolve() if args.output_duplicates else (feixi_dir / "\u6807\u8bc6\u7801\u91cd\u590d\u6e05\u5355.csv")
    non_g_csv = Path(args.output_non_g_report).resolve() if args.output_non_g_report else (feixi_dir / "\u975eg\u89c4\u683c\u5bf9\u7167.csv")

    default_min = float(args.default_min_stock)
    default_loc = args.default_location

    try:
        inv_rows = read_inventory_rows(inventory_file)
        source_rows = load_source_rows(source_file)
        source_profile_map = build_source_profile_map(source_rows)
        (
            final_rows,
            code_counter,
            matched_source_count,
            source_only_forced_disabled_count,
            inv_non_positive_forced_disabled_count,
        ) = build_final_rows(
            inv_rows,
            source_profile_map=source_profile_map,
            default_min=default_min,
            default_loc=default_loc,
        )
        write_duplicate_csv(dup_csv, inv_rows, code_counter)
        write_template_rows(template_file, final_out, final_rows)

        backup_rows = build_backup_rows(source_rows, default_min=default_min, default_loc=default_loc)
        write_template_rows(template_file, backup_out, backup_rows)
        write_non_g_report(non_g_csv, inv_rows, source_rows)
    except Exception as exc:  # noqa: BLE001
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1

    dup_count = sum(1 for c, n in code_counter.items() if n > 1)
    dup_rows = sum(n for _, n in code_counter.items() if n > 1)
    non_g = [r for r in inv_rows if r.spec and "g" not in r.spec.lower()]

    print("[OK] Feixi conversion complete")
    print(f"inventory_file={inventory_file}")
    print(f"source_file={source_file}")
    print(f"template_file={template_file}")
    print(f"inventory_rows={len(inv_rows)}")
    print(f"inventory_unique_codes={len(code_counter)}")
    print(f"output_unique_codes={len(final_rows)}")
    print(f"matched_from_source={matched_source_count}")
    print(f"defaulted_not_found={len(final_rows)-matched_source_count}")
    print(f"source_only_forced_disabled={source_only_forced_disabled_count}")
    print(f"inventory_non_positive_forced_disabled={inv_non_positive_forced_disabled_count}")
    print(f"duplicate_code_count={dup_count}")
    print(f"duplicate_rows={dup_rows}")
    print(f"non_g_spec_rows={len(non_g)}")
    print(f"final_output={final_out}")
    print(f"backup_output={backup_out}")
    print(f"duplicate_csv={dup_csv}")
    print(f"non_g_report_csv={non_g_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
