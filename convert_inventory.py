from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

CN_CODE = "\u996e\u7247\u7f16\u7801"
CN_BATCH = "\u6279\u6b21"
CN_STOCK = "\u5e93\u5b58"
CN_STATUS = "\u72b6\u6001"
CN_ENABLED = "\u662f\u5426\u542f\u7528"
CN_LOCATION = "\u8d27\u4f4d\u7f16\u53f7"
CN_LOCATION_ALT = "\u8d27\u4f4d\u7f16\u7801"
CN_MIN_STOCK = "\u5e93\u5b58\u4e0b\u9650\u503c"

CN_STATUS_ON = "\u542f\u7528"
CN_STATUS_OFF = "\u7981\u7528"
CN_YES = "\u662f"
CN_NO = "\u5426"

INVENTORY_REQUIRED_HEADERS = (CN_CODE, CN_STOCK, CN_STATUS)
TEMPLATE_REQUIRED_HEADERS = (CN_CODE, CN_ENABLED, CN_LOCATION, CN_STOCK, CN_MIN_STOCK)

# Legacy inventory layout fallback when header row is missing.
FALLBACK_CODE_COL = 2
FALLBACK_BATCH_COL = 3
FALLBACK_STOCK_COL = 11
FALLBACK_STATUS_COL = 12
RE_NUMERIC_TEXT = re.compile(r"^[+-]?(?:\d+(?:\.\d+)?|\.\d+)(?:[eE][+-]?\d+)?$")


@dataclass
class InventoryRecord:
    row_num: int
    code: str
    batch: str
    stock: float
    status: str


@dataclass
class OutputRow:
    code: str
    enabled: str
    location: str
    stock: float | int
    min_stock: float | int


@dataclass
class SourceProfile:
    location: str | None
    min_stock: float | None


@dataclass
class InventoryLayout:
    start_row: int
    code_col: int
    batch_col: int | None
    stock_col: int
    status_col: int


def decode_maybe_gbk(value: object) -> object:
    if isinstance(value, str):
        try:
            return value.encode("latin1").decode("gbk")
        except Exception:
            return value
    return value


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = decode_maybe_gbk(value)
    return str(text).strip()


def code_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return _normalize_numeric_string(str(value))

    text = normalize_text(value)
    if text == "":
        return ""

    plain = text.replace(",", "")
    # Preserve user-provided leading-zero pure digit codes.
    if re.match(r"^[+-]?0\d+$", plain):
        return text
    if not RE_NUMERIC_TEXT.match(plain):
        return text
    return _normalize_numeric_string(plain)


def _parse_decimal(raw: str) -> Decimal | None:
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _normalize_numeric_string(raw: str) -> str:
    dec = _parse_decimal(raw)
    if dec is None:
        return raw
    if dec == dec.to_integral_value():
        return format(dec.quantize(Decimal("1")), "f")
    normalized = dec.normalize()
    text = format(normalized, "f")
    return text.rstrip("0").rstrip(".")


def to_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    raw = normalize_text(value).replace(",", "")
    if raw == "":
        return 0.0
    dec = _parse_decimal(raw)
    if dec is None:
        return 0.0
    return float(dec)


def to_optional_number(value: object) -> float | None:
    text = normalize_text(value)
    if text == "":
        return None
    dec = _parse_decimal(text.replace(",", ""))
    if dec is None:
        return None
    return float(dec)


def to_excel_number(value: float) -> float | int:
    dec = Decimal(str(value))
    rounded_int = dec.to_integral_value()
    if dec == rounded_int:
        return int(rounded_int)
    quantized = dec.quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)
    return float(quantized)


def read_header_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = normalize_text(ws.cell(row=1, column=col).value)
        if key and key not in header_map:
            header_map[key] = col
    return header_map


def ensure_headers(header_map: dict[str, int], required: tuple[str, ...], file_label: str) -> None:
    missing = [h for h in required if h not in header_map]
    if missing:
        raise ValueError(f"{file_label} 缺少必需列: {', '.join(missing)}")


def detect_inventory_layout(ws: openpyxl.worksheet.worksheet.Worksheet) -> InventoryLayout:
    header_map = read_header_map(ws)
    has_header = all(k in header_map for k in INVENTORY_REQUIRED_HEADERS)

    if has_header:
        batch_col = header_map.get(CN_BATCH)
        return InventoryLayout(
            start_row=2,
            code_col=header_map[CN_CODE],
            batch_col=batch_col,
            stock_col=header_map[CN_STOCK],
            status_col=header_map[CN_STATUS],
        )

    # Headerless fallback: use known legacy column positions.
    if ws.max_column < FALLBACK_STATUS_COL:
        raise ValueError(
            f"库存文件既没有表头，也不符合旧版列位结构（至少需要 {FALLBACK_STATUS_COL} 列）"
        )

    return InventoryLayout(
        start_row=1,
        code_col=FALLBACK_CODE_COL,
        batch_col=FALLBACK_BATCH_COL if ws.max_column >= FALLBACK_BATCH_COL else None,
        stock_col=FALLBACK_STOCK_COL,
        status_col=FALLBACK_STATUS_COL,
    )


def load_inventory_records(inventory_path: Path) -> list[InventoryRecord]:
    wb = openpyxl.load_workbook(inventory_path, data_only=True)
    ws = wb.worksheets[0]
    layout = detect_inventory_layout(ws)

    records: list[InventoryRecord] = []
    for row in range(layout.start_row, ws.max_row + 1):
        code = code_to_text(ws.cell(row=row, column=layout.code_col).value)
        if not code:
            continue

        batch = ""
        if layout.batch_col is not None:
            batch = normalize_text(ws.cell(row=row, column=layout.batch_col).value)

        stock = to_number(ws.cell(row=row, column=layout.stock_col).value)
        status = normalize_text(ws.cell(row=row, column=layout.status_col).value)
        records.append(
            InventoryRecord(
                row_num=row,
                code=code,
                batch=batch,
                stock=stock,
                status=status,
            )
        )

    if not records:
        raise ValueError("库存文件未读取到有效数据行")
    return records


def map_status(status_set: set[str]) -> str:
    if CN_STATUS_ON in status_set:
        return CN_YES
    if CN_STATUS_OFF in status_set:
        return CN_NO
    if CN_YES in status_set:
        return CN_YES
    if CN_NO in status_set:
        return CN_NO
    return ""


def load_workbook_any_excel(path: Path):
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except InvalidFileException:
        with path.open("rb") as fh:
            return openpyxl.load_workbook(fh, data_only=True)


def merge_source_profile(
    profiles: dict[str, SourceProfile],
    code: str,
    location: str | None,
    min_stock: float | None,
) -> None:
    existing = profiles.get(code)
    if existing is None:
        profiles[code] = SourceProfile(location=location, min_stock=min_stock)
        return

    if (existing.location is None or existing.location == "") and location:
        existing.location = location
    if existing.min_stock is None and min_stock is not None:
        existing.min_stock = min_stock


def load_source_profiles_xlsx(source_path: Path) -> dict[str, SourceProfile]:
    wb = load_workbook_any_excel(source_path)
    ws = wb.worksheets[0]
    header = read_header_map(ws)

    if CN_CODE not in header:
        raise ValueError(f"源文件缺少列: {CN_CODE}")

    location_col = header.get(CN_LOCATION) or header.get(CN_LOCATION_ALT)
    min_col = header.get(CN_MIN_STOCK)
    code_col = header[CN_CODE]

    profiles: dict[str, SourceProfile] = {}
    for r in range(2, ws.max_row + 1):
        code = normalize_text(ws.cell(r, code_col).value)
        if not code:
            continue
        location = normalize_text(ws.cell(r, location_col).value) if location_col else ""
        min_stock = to_optional_number(ws.cell(r, min_col).value) if min_col else None
        merge_source_profile(
            profiles,
            code=code,
            location=location if location else None,
            min_stock=min_stock,
        )
    return profiles


def load_source_profiles_xls(source_path: Path) -> dict[str, SourceProfile]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("读取 .xls 需要 xlrd，请先执行 uv sync 安装依赖。") from exc

    try:
        wb = xlrd.open_workbook(str(source_path))
    except Exception as exc:  # noqa: BLE001
        if "xlsx file; not supported" in str(exc).lower():
            return load_source_profiles_xlsx(source_path)
        raise

    ws = wb.sheet_by_index(0)
    headers = [normalize_text(v) for v in ws.row_values(0)]
    header_map = {name: idx for idx, name in enumerate(headers) if name}

    if CN_CODE not in header_map:
        raise ValueError(f"源文件缺少列: {CN_CODE}")

    location_col = header_map.get(CN_LOCATION)
    if location_col is None:
        location_col = header_map.get(CN_LOCATION_ALT)
    min_col = header_map.get(CN_MIN_STOCK)
    code_col = header_map[CN_CODE]

    profiles: dict[str, SourceProfile] = {}
    for r in range(1, ws.nrows):
        row = ws.row_values(r)
        code = normalize_text(row[code_col] if code_col < len(row) else "")
        if not code:
            continue
        location = normalize_text(row[location_col]) if location_col is not None and location_col < len(row) else ""
        min_stock = to_optional_number(row[min_col]) if min_col is not None and min_col < len(row) else None
        merge_source_profile(
            profiles,
            code=code,
            location=location if location else None,
            min_stock=min_stock,
        )
    return profiles


def load_source_profiles(source_path: Path | None) -> dict[str, SourceProfile]:
    if source_path is None:
        return {}
    if not source_path.exists():
        raise FileNotFoundError(f"源文件不存在: {source_path}")

    suffix = source_path.suffix.lower()
    if suffix == ".xlsx":
        return load_source_profiles_xlsx(source_path)
    if suffix == ".xls":
        return load_source_profiles_xls(source_path)
    raise ValueError(f"不支持的源文件格式: {source_path.suffix}")


def aggregate_records(
    records: list[InventoryRecord],
    source_profiles: dict[str, SourceProfile],
    location_default: str,
    min_stock_default: float | int,
) -> tuple[list[OutputRow], Counter[tuple[str, str]], dict[str, set[str]], int]:
    pair_counter: Counter[tuple[str, str]] = Counter((r.code, r.batch) for r in records)
    code_stock: dict[str, float] = defaultdict(float)
    code_status: dict[str, set[str]] = defaultdict(set)
    code_order: list[str] = []
    seen_codes: set[str] = set()

    for r in records:
        code_stock[r.code] += r.stock
        if r.status:
            code_status[r.code].add(r.status)
        if r.code not in seen_codes:
            seen_codes.add(r.code)
            code_order.append(r.code)

    rows: list[OutputRow] = []
    matched_source_count = 0
    for code in code_order:
        profile = source_profiles.get(code)
        if profile is not None:
            matched_source_count += 1
            location = profile.location if profile.location else location_default
            min_stock = profile.min_stock if profile.min_stock is not None else float(min_stock_default)
        else:
            location = location_default
            min_stock = float(min_stock_default)

        rows.append(
            OutputRow(
                code=code,
                enabled=map_status(code_status[code]),
                location=location,
                stock=to_excel_number(code_stock[code]),
                min_stock=to_excel_number(min_stock),
            )
        )

    return rows, pair_counter, code_status, matched_source_count


def write_template(
    template_path: Path,
    output_path: Path,
    rows: list[OutputRow],
    sort_desc: bool,
) -> None:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]
    header_map = read_header_map(ws)
    ensure_headers(header_map, TEMPLATE_REQUIRED_HEADERS, f"模板文件 {template_path}")

    if sort_desc:
        rows = sorted(rows, key=lambda x: float(x.stock), reverse=True)

    code_col = header_map[CN_CODE]
    enabled_col = header_map[CN_ENABLED]
    location_col = header_map[CN_LOCATION]
    stock_col = header_map[CN_STOCK]
    min_stock_col = header_map[CN_MIN_STOCK]

    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    for row_idx, row_data in enumerate(rows, start=2):
        code_cell = ws.cell(row=row_idx, column=code_col, value=row_data.code)
        enabled_cell = ws.cell(row=row_idx, column=enabled_col, value=row_data.enabled)
        location_cell = ws.cell(row=row_idx, column=location_col, value=row_data.location)
        stock_cell = ws.cell(row=row_idx, column=stock_col, value=row_data.stock)
        min_stock_cell = ws.cell(row=row_idx, column=min_stock_col, value=row_data.min_stock)

        # Keep identifiers/text fields from being displayed in scientific notation.
        code_cell.number_format = "@"
        enabled_cell.number_format = "@"
        location_cell.number_format = "@"
        # Keep numeric display stable and avoid Excel scientific-format auto-switch.
        stock_cell.number_format = "0.######"
        min_stock_cell.number_format = "0.######"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_reports(
    report_dir: Path,
    records: list[InventoryRecord],
    pair_counter: Counter[tuple[str, str]],
    code_status: dict[str, set[str]],
) -> tuple[Path, Path, Path]:
    report_dir.mkdir(parents=True, exist_ok=True)

    summary_file = report_dir / "duplicate_code_batch_summary.csv"
    details_file = report_dir / "duplicate_code_batch_details.csv"
    conflict_file = report_dir / "code_status_conflicts.csv"

    duplicate_pairs = sorted((k for k, count in pair_counter.items() if count > 1))

    with summary_file.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["code", "batch", "pair_count"])
        for code, batch in duplicate_pairs:
            writer.writerow([code, batch, pair_counter[(code, batch)]])

    with details_file.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["code", "batch", "pair_count", "source_row", "stock", "status"])
        for record in records:
            key = (record.code, record.batch)
            if key in duplicate_pairs:
                writer.writerow(
                    [
                        record.code,
                        record.batch,
                        pair_counter[key],
                        record.row_num,
                        record.stock,
                        record.status,
                    ]
                )

    with conflict_file.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["code", "statuses"])
        for code in sorted(code_status.keys()):
            statuses = {s for s in code_status[code] if s}
            if len(statuses) > 1:
                writer.writerow([code, "|".join(sorted(statuses))])

    return summary_file, details_file, conflict_file


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="将药品库存转换为饮片货位导入模板（按饮片编码聚合库存）"
    )
    parser.add_argument("--inventory", required=True, help="库存文件路径（xlsx）")
    parser.add_argument("--template", required=True, help="导入模板路径（xlsx）")
    parser.add_argument(
        "--source-for-match",
        help="配方间源数据路径（xls/xlsx），用于匹配最终文件的货位编号和库存下限值",
    )
    parser.add_argument(
        "--output",
        help="输出文件路径（xlsx）。不传时默认在模板同目录生成 *_generated.xlsx",
    )
    parser.add_argument(
        "--location",
        default="Z999",
        help="货位编号默认值，默认 Z999",
    )
    parser.add_argument(
        "--min-stock",
        type=float,
        default=500,
        help="库存下限值默认值，默认 500",
    )
    parser.add_argument(
        "--no-sort",
        action="store_true",
        help="不按库存倒序排序（默认按库存从大到小排序）",
    )
    parser.add_argument(
        "--report-dir",
        default=".",
        help="报告文件输出目录，默认当前目录",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()

    inventory_path = Path(args.inventory).expanduser().resolve()
    template_path = Path(args.template).expanduser().resolve()
    source_match_path = Path(args.source_for_match).expanduser().resolve() if args.source_for_match else None

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        output_path = template_path.with_name(f"{template_path.stem}_generated{template_path.suffix}")

    report_dir = Path(args.report_dir).expanduser().resolve()

    if not inventory_path.exists():
        print(f"[ERROR] 库存文件不存在: {inventory_path}", file=sys.stderr)
        return 1
    if not template_path.exists():
        print(f"[ERROR] 模板文件不存在: {template_path}", file=sys.stderr)
        return 1

    min_stock_value: float | int = args.min_stock
    if abs(args.min_stock - round(args.min_stock)) < 1e-9:
        min_stock_value = int(round(args.min_stock))

    try:
        records = load_inventory_records(inventory_path)
        source_profiles = load_source_profiles(source_match_path)
        rows, pair_counter, code_status, matched_source_count = aggregate_records(
            records=records,
            source_profiles=source_profiles,
            location_default=args.location,
            min_stock_default=min_stock_value,
        )
        write_template(
            template_path=template_path,
            output_path=output_path,
            rows=rows,
            sort_desc=not args.no_sort,
        )
        summary_file, details_file, conflict_file = write_reports(
            report_dir=report_dir,
            records=records,
            pair_counter=pair_counter,
            code_status=code_status,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"[ERROR] 处理失败: {exc}", file=sys.stderr)
        return 1

    duplicate_pair_count = sum(1 for _, count in pair_counter.items() if count > 1)
    duplicate_pair_rows = sum(count for _, count in pair_counter.items() if count > 1)

    print("[OK] 转换完成")
    print(f"inventory: {inventory_path}")
    print(f"template:  {template_path}")
    print(f"source_for_match: {source_match_path if source_match_path else ''}")
    print(f"output:    {output_path}")
    print(f"source_rows: {len(records)}")
    print(f"unique_codes: {len(rows)}")
    print(f"matched_from_source: {matched_source_count}")
    print(f"defaulted_not_found: {len(rows)-matched_source_count}")
    print(f"duplicate_code_batch_count: {duplicate_pair_count}")
    print(f"duplicate_code_batch_rows:  {duplicate_pair_rows}")
    print(f"report_summary: {summary_file}")
    print(f"report_details: {details_file}")
    print(f"report_status_conflicts: {conflict_file}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
